import sqlite3
import os
import csv
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import List, Optional

DB_FILE = "coworking.db"

TURNOS = {
    "M": "Matutino",
    "V": "Vespertino",
    "N": "Nocturno",
}

@dataclass
class Cliente:
    id: str
    nombres: str
    apellidos: str

@dataclass
class Sala:
    id: str
    nombre: str
    cupo: int

@dataclass
class Reservacion:
    folio: int
    evento: str
    id_cliente: str
    id_sala: str
    fecha: datetime
    turno: str
    estado: str

class BaseDatos:
    def _init_(self, db_file: str = DB_FILE):
        self.db_file = db_file
        self.conn = sqlite3.connect(
            db_file,
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
        )
        self._inicializar()

    def _inicializar(self):
        cursor = self.conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id TEXT PRIMARY KEY,
                nombres TEXT NOT NULL,
                apellidos TEXT NOT NULL
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS salas (
                id TEXT PRIMARY KEY,
                nombre TEXT NOT NULL,
                cupo INTEGER NOT NULL
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS reservaciones (
                folio INTEGER PRIMARY KEY AUTOINCREMENT,
                evento TEXT NOT NULL,
                id_cliente TEXT NOT NULL,
                id_sala TEXT NOT NULL,
                fecha timestamp NOT NULL,
                turno TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'activa',
                FOREIGN KEY (id_cliente) REFERENCES clientes(id),
                FOREIGN KEY (id_sala) REFERENCES salas(id)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS contadores (
                tipo TEXT PRIMARY KEY,
                valor INTEGER NOT NULL
            )
        """)

        for tipo in ["C", "S"]:
            cursor.execute("INSERT OR IGNORE INTO contadores (tipo, valor) VALUES (?, 0)", (tipo,))

        self.conn.commit()

    def _nuevo_id(self, prefijo: str) -> str:
        cursor = self.conn.cursor()
        cursor.execute("UPDATE contadores SET valor = valor + 1 WHERE tipo = ?", (prefijo,))
        cursor.execute("SELECT valor FROM contadores WHERE tipo = ?", (prefijo,))
        row = cursor.fetchone()
        valor = row[0] if row else 0
        self.conn.commit()
        return f"{prefijo}{valor:04d}"

    def registrar_cliente(self, nombres: str, apellidos: str) -> Cliente:
        nombres = nombres.strip()
        apellidos = apellidos.strip()
        if not nombres or not apellidos:
            raise ValueError("Nombres y apellidos no pueden estar vacíos.")
        cursor = self.conn.cursor()
        cursor.execute("SELECT id FROM clientes WHERE LOWER(nombres) = ? AND LOWER(apellidos) = ?", (nombres.lower(), apellidos.lower()))
        if cursor.fetchone():
            raise ValueError("El cliente ya existe.")
        cid = self._nuevo_id("C")
        cursor.execute("INSERT INTO clientes (id, nombres, apellidos) VALUES (?, ?, ?)", (cid, nombres, apellidos))
        self.conn.commit()
        return Cliente(id=cid, nombres=nombres, apellidos=apellidos)

    def listar_clientes_ordenados(self) -> List[Cliente]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, nombres, apellidos FROM clientes ORDER BY apellidos, nombres")
        return [Cliente(*row) for row in cursor.fetchall()]

    def registrar_sala(self, nombre: str, cupo: int) -> Sala:
        nombre = nombre.strip()
        if not nombre:
            raise ValueError("El nombre de la sala no puede estar vacío.")
        if cupo <= 0:
            raise ValueError("El cupo debe ser mayor que 0.")
        cursor = self.conn.cursor()
        cursor.execute("SELECT id FROM salas WHERE LOWER(nombre) = ?", (nombre.lower(),))
        if cursor.fetchone():
            raise ValueError("Ya existe una sala con ese nombre.")
        sid = self._nuevo_id("S")
        cursor.execute("INSERT INTO salas (id, nombre, cupo) VALUES (?, ?, ?)", (sid, nombre, cupo))
        self.conn.commit()
        return Sala(id=sid, nombre=nombre, cupo=cupo)

    def obtener_sala(self, id_sala: str) -> Optional[Sala]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, nombre, cupo FROM salas WHERE id = ?", (id_sala,))
        row = cursor.fetchone()
        return Sala(*row) if row else None

    def obtener_cliente(self, id_cliente: str) -> Optional[Cliente]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, nombres, apellidos FROM clientes WHERE id = ?", (id_cliente,))
        row = cursor.fetchone()
        return Cliente(*row) if row else None

    def salas_disponibles(self, fecha_dt: datetime, turno: str) -> List[Sala]:
        cursor = self.conn.cursor()
        fecha_buscar = fecha_dt.date()
        cursor.execute("""
            SELECT s.id, s.nombre, s.cupo FROM salas s
            WHERE s.id NOT IN (
                SELECT id_sala FROM reservaciones 
                WHERE DATE(fecha) = :fecha AND turno = :turno AND estado = 'activa'
            )
        """, {"fecha": fecha_buscar, "turno": turno})
        return [Sala(*row) for row in cursor.fetchall()]

    def registrar_reserva(self, evento: str, id_cliente: str, id_sala: str, fecha_dt: datetime, turno: str) -> Reservacion:
        evento = (evento or "").strip()
        if not evento:
            raise ValueError("El nombre del evento no puede estar vacío.")
        cursor = self.conn.cursor()
        cursor.execute("SELECT id FROM clientes WHERE id = ?", (id_cliente,))
        if not cursor.fetchone():
            raise ValueError("Cliente no encontrado.")
        cursor.execute("SELECT id FROM salas WHERE id = ?", (id_sala,))
        if not cursor.fetchone():
            raise ValueError("Sala no encontrada.")
        if turno not in TURNOS:
            raise ValueError("Turno inválido.")
        fecha_buscar = fecha_dt.date()
        cursor.execute("""
            SELECT folio FROM reservaciones 
            WHERE id_sala = :sala AND DATE(fecha) = :fecha AND turno = :turno AND estado = 'activa'
        """, {"sala": id_sala, "fecha": fecha_buscar, "turno": turno})
        if cursor.fetchone():
            raise ValueError("Ya existe una reservación activa en esa sala para esa fecha y turno.")
        cursor.execute("""
            INSERT INTO reservaciones (evento, id_cliente, id_sala, fecha, turno, estado)
            VALUES (?, ?, ?, ?, ?, 'activa')
        """, (evento, id_cliente, id_sala, fecha_dt, turno))
        self.conn.commit()
        folio = cursor.lastrowid
        return Reservacion(folio=folio, evento=evento, id_cliente=id_cliente, id_sala=id_sala, fecha=fecha_dt, turno=turno, estado='activa')

    def reservas_en_rango(self, desde_dt: datetime, hasta_dt: datetime) -> List[Reservacion]:
        cursor = self.conn.cursor()
        desde_buscar = desde_dt.date()
        hasta_buscar = hasta_dt.date()
        cursor.execute("""
            SELECT folio, evento, id_cliente, id_sala, fecha, turno, estado
            FROM reservaciones
            WHERE DATE(fecha) BETWEEN :desde AND :hasta AND estado = 'activa'
            ORDER BY fecha, folio
        """, {"desde": desde_buscar, "hasta": hasta_buscar})
        return [Reservacion(*row) for row in cursor.fetchall()]

    def editar_nombre_evento(self, folio: int, nuevo_nombre: str) -> Reservacion:
        nuevo_nombre = (nuevo_nombre or "").strip()
        if not nuevo_nombre:
            raise ValueError("El nuevo nombre no puede estar vacío.")
        cursor = self.conn.cursor()
        cursor.execute("SELECT folio, estado FROM reservaciones WHERE folio = ?", (folio,))
        row = cursor.fetchone()
        if not row:
            raise ValueError("Folio no encontrado.")
        if row[1] == 'cancelada':
            raise ValueError("No se puede editar una reservación cancelada.")
        cursor.execute("UPDATE reservaciones SET evento = ? WHERE folio = ?", (nuevo_nombre, folio))
        self.conn.commit()
        cursor.execute("SELECT folio, evento, id_cliente, id_sala, fecha, turno, estado FROM reservaciones WHERE folio = ?", (folio,))
        return Reservacion(*cursor.fetchone())

    def reservas_por_fecha(self, fecha_dt: datetime) -> List[Reservacion]:
        cursor = self.conn.cursor()
        fecha_buscar = fecha_dt.date()
        cursor.execute("""
            SELECT folio, evento, id_cliente, id_sala, fecha, turno, estado
            FROM reservaciones
            WHERE DATE(fecha) = :fecha AND estado = 'activa'
            ORDER BY turno, folio
        """, {"fecha": fecha_buscar})
        return [Reservacion(*row) for row in cursor.fetchall()]

    def cancelar_reservacion(self, folio: int) -> Reservacion:
        cursor = self.conn.cursor()
        cursor.execute("SELECT folio, evento, id_cliente, id_sala, fecha, turno, estado FROM reservaciones WHERE folio = ?", (folio,))
        row = cursor.fetchone()
        if not row:
            raise ValueError("Folio no encontrado.")
        reserva = Reservacion(*row)
        if reserva.estado == 'cancelada':
            raise ValueError("Esta reservación ya está cancelada.")
        fecha_reserva = reserva.fecha.date()
        dias_anticipacion = (fecha_reserva - date.today()).days
        if dias_anticipacion < 2:
            raise ValueError(f"Solo puede cancelar con al menos 2 días de anticipación. Días restantes: {dias_anticipacion}")
        cursor.execute("UPDATE reservaciones SET estado = 'cancelada' WHERE folio = ?", (folio,))
        self.conn.commit()
        reserva.estado = 'cancelada'
        return reserva

    def cerrar(self):
        self.conn.close()

# ---------------------------
# Utilidades y UI
# ---------------------------
def linea(ancho: int = 80, char: str = "-") -> str:
    return char * ancho

def tabla(headers: List[str], filas: List[List[str]]) -> str:
    if not filas:
        return ""
    col_widths = [max(len(str(h)), *(len(str(f[i])) for f in filas)) for i, h in enumerate(headers)]
    def fmt_row(row): return " | ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(headers)))
    out = [fmt_row(headers), "-+-".join("-" * w for w in col_widths)]
    for f in filas: out.append(fmt_row(f))
    return "\n".join(out)

def input_no_vacio(prompt: str) -> str:
    while True:
        val = input(prompt).strip()
        if val and not val.isspace():
            return val
        print("⚠ No puede estar vacío ni contener solo espacios.")

def input_entero(prompt: str, minimo: Optional[int] = None) -> int:
    while True:
        try:
            v = int(input(prompt).strip())
            if minimo is not None and v < minimo:
                print(f"⚠ Debe ser ≥ {minimo}.")
                continue
            return v
        except ValueError:
            print("⚠ Ingrese un número válido.")

def input_fecha(prompt: str, permitir_vacio: bool = False) -> datetime:
    while True:
        s = input(prompt).strip()
        if permitir_vacio and not s:
            return datetime.combine(date.today(), datetime.min.time())
        try:
            fecha = datetime.strptime(s, "%m-%d-%Y")
            return fecha
        except ValueError:
            print("⚠ Formato inválido. Use mm-dd-aaaa (ejemplo: 12-25-2025).")

def validar_fecha_reservacion(fecha_dt: datetime) -> bool:
    if fecha_dt.date() < (date.today() + timedelta(days=2)):
        raise ValueError("La reservación debe ser al menos 2 días después de hoy.")
    return True

def es_domingo(fecha_dt: datetime) -> bool:
    return fecha_dt.weekday() == 6

def obtener_lunes_siguiente(fecha_dt: datetime) -> datetime:
    dias_hasta_lunes = (7 - fecha_dt.weekday()) % 7
    if dias_hasta_lunes == 0:
        dias_hasta_lunes = 1
    lunes = fecha_dt + timedelta(days=dias_hasta_lunes)
    return lunes

def fecha_a_str(fecha_dt: datetime) -> str:
    return fecha_dt.strftime("%m-%d-%Y")

def pausar():
    input("\n[Presione ENTER para continuar...]")

# ---------------------------
# Opciones del menú
# ---------------------------
def opcion_registrar_reserva(db: BaseDatos):
    print(linea())
    print("REGISTRAR RESERVACIÓN DE SALA")
    print(linea())

    clientes = db.listar_clientes_ordenados()
    if not clientes:
        print("⚠ No hay clientes registrados. Registre un cliente primero.")
        pausar()
        return

    cursor = db.conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM salas")
    if cursor.fetchone()[0] == 0:
        print("⚠ No hay salas registradas. Registre una sala primero.")
        pausar()
        return

    print("\nClientes registrados (ordenados alfabéticamente):")
    print(linea())
    filas = [[c.id, f"{c.apellidos}, {c.nombres}"] for c in clientes]
    print(tabla(["Clave Cliente", "Apellidos, Nombres"], filas))
    print(linea())

    while True:
        id_cliente = input("\nClave del cliente (o 'CANCELAR' para salir): ").strip()
        if id_cliente.upper() == 'CANCELAR':
            print("Operación cancelada.")
            pausar()
            return
        if db.obtener_cliente(id_cliente):
            break
        print("\n⚠ La clave seleccionada no existe.")
        print("\nClientes registrados:")
        print(linea())
        print(tabla(["Clave Cliente", "Apellidos, Nombres"], filas))
        print(linea())

    print(f"\nFecha actual del sistema: {date.today().strftime('%m-%d-%Y')}")
    print(f"La fecha debe ser al menos: {(date.today() + timedelta(days=2)).strftime('%m-%d-%Y')}")

    while True:
        fecha_dt = input_fecha("\nFecha de reservación (mm-dd-aaaa): ")
        try:
            validar_fecha_reservacion(fecha_dt)
            if es_domingo(fecha_dt):
                lunes = obtener_lunes_siguiente(fecha_dt)
                print(f"\n⚠ No se pueden hacer reservaciones para domingos.")
                print(f"   Se propone el lunes siguiente: {fecha_a_str(lunes)}")
                acepta = input("¿Acepta esta fecha? (S/N): ").strip().upper()
                if acepta == 'S':
                    fecha_dt = lunes
                    break
                else:
                    print("Por favor, especifique otra fecha.")
                    continue
            break
        except ValueError as e:
            print(f"⚠ {e}")

    print("\nSeleccione turno:")
    print("  M) Matutino")
    print("  V) Vespertino")
    print("  N) Nocturno")

    while True:
        turno = input("Turno [M/V/N]: ").strip().upper()
        if turno in TURNOS:
            break
        print("⚠ Turno inválido. Use M, V o N.")

    try:
        salas_disp = db.salas_disponibles(fecha_dt, turno)

        if not salas_disp:
            print(f"\n⚠ No hay salas disponibles para el turno {TURNOS[turno]} el {fecha_a_str(fecha_dt)}")
            pausar()
            return

        print(f"\nSalas disponibles para {TURNOS[turno]} el {fecha_a_str(fecha_dt)}:")
        print(linea())
        filas_salas = [[s.id, s.nombre, str(s.cupo)] for s in salas_disp]
        print(tabla(["Clave Sala", "Nombre", "Cupo"], filas_salas))
        print(linea())

        while True:
            id_sala = input("\nClave de la sala: ").strip()
            if id_sala in [s.id for s in salas_disp]:
                break
            print("⚠ Sala no disponible o no válida.")

        evento = input_no_vacio("\nNombre del evento: ")

        reserva = db.registrar_reserva(evento, id_cliente, id_sala, fecha_dt, turno)

        cliente = db.obtener_cliente(id_cliente)
        sala = db.obtener_sala(id_sala)

        print("\n" + linea())
        print("✓ RESERVACIÓN REGISTRADA EXITOSAMENTE")
        print(linea())
        print(f"  Folio:   {reserva.folio}")
        print(f"  Evento:  {reserva.evento}")
        print(f"  Cliente: {cliente.apellidos}, {cliente.nombres}")
        print(f"  Sala:    {sala.nombre}")
        print(f"  Fecha:   {fecha_a_str(reserva.fecha)}")
        print(f"  Turno:   {TURNOS[turno]}")
        print(linea())

    except ValueError as e:
        print(f"\n✗ Error: {e}")
    except sqlite3.Error as e:
        print(f"\n✗ Error de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    pausar()

def opcion_registrar_cliente(db: BaseDatos):
    print(linea())
    print("REGISTRAR NUEVO CLIENTE")
    print(linea())
    try:
        nombres = input_no_vacio("Nombres: ")
        apellidos = input_no_vacio("Apellidos: ")
        cliente = db.registrar_cliente(nombres, apellidos)
        print(f"\n✓ Cliente registrado exitosamente")
        print(f"  Clave generada: {cliente.id}")
        print(f"  Nombre: {cliente.nombres} {cliente.apellidos}")
    except ValueError as e:
        print(f"\n✗ Error: {e}")
    except sqlite3.Error as e:
        print(f"\n✗ Error de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    pausar()

def opcion_registrar_sala(db: BaseDatos):
    print(linea())
    print("REGISTRAR NUEVA SALA")
    print(linea())
    try:
        nombre = input_no_vacio("Nombre de la sala: ")
        cupo = input_entero("Cupo de la sala: ", minimo=1)
        sala = db.registrar_sala(nombre, cupo)
        print(f"\n✓ Sala registrada exitosamente")
        print(f"  Clave generada: {sala.id}")
        print(f"  Nombre: {sala.nombre}")
        print(f"  Cupo: {sala.cupo}")
    except ValueError as e:
        print(f"\n✗ Error: {e}")
    except sqlite3.Error as e:
        print(f"\n✗ Error de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    pausar()

def opcion_editar_evento(db: BaseDatos):
    print(linea())
    print("EDITAR NOMBRE DE EVENTO DE UNA RESERVACIÓN")
    print(linea())

    fecha_desde = input_fecha("Fecha inicial del rango (mm-dd-aaaa): ")
    fecha_hasta = input_fecha("Fecha final del rango (mm-dd-aaaa): ")

    try:
        reservas = db.reservas_en_rango(fecha_desde, fecha_hasta)

        if not reservas:
            print(f"\n⚠ No hay reservaciones en el rango {fecha_a_str(fecha_desde)} a {fecha_a_str(fecha_hasta)}")
            pausar()
            return

        print(f"\nEventos registrados del {fecha_a_str(fecha_desde)} al {fecha_a_str(fecha_hasta)}:")
        print(linea())
        filas = [[str(r.folio), r.evento, fecha_a_str(r.fecha)] for r in reservas]
        print(tabla(["Folio", "Nombre del Evento", "Fecha"], filas))
        print(linea())

        while True:
            folio_str = input("\nFolio del evento a modificar (o 'CANCELAR' para salir): ").strip()
            if folio_str.upper() == 'CANCELAR':
                print("Operación de modificación cancelada.")
                pausar()
                return
            try:
                folio = int(folio_str)
                if folio in [r.folio for r in reservas]:
                    break
            except ValueError:
                pass
            print("\n⚠ El folio indicado no pertenece a este rango.")
            print("\nEventos disponibles:")
            print(linea())
            print(tabla(["Folio", "Nombre del Evento", "Fecha"], filas))
            print(linea())

        nuevo_nombre = input_no_vacio("\nNuevo nombre del evento: ")
        db.editar_nombre_evento(folio, nuevo_nombre)

        print("\n" + linea())
        print("✓ Nombre del evento actualizado exitosamente")
        print(linea())

    except ValueError as e:
        print(f"\n✗ Error: {e}")
    except sqlite3.Error as e:
        print(f"\n✗ Error de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    pausar()

def opcion_consultar_por_fecha(db: BaseDatos):
    print(linea())
    print("CONSULTAR RESERVACIONES EXISTENTES PARA UNA FECHA ESPECÍFICA")
    print(linea())
    print("(Deje vacío para usar la fecha actual del sistema)")
    fecha_dt = input_fecha("Fecha a consultar (mm-dd-aaaa): ", permitir_vacio=True)

    try:
        lista = db.reservas_por_fecha(fecha_dt)
    except sqlite3.Error as e:
        print(f"✗ Error de base de datos: {e}")
        pausar()
        return
    except Exception:
        print(f"✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
        pausar()
        return

    if not lista:
        print(f"\nNo hay reservaciones para la fecha {fecha_a_str(fecha_dt)}.")
        pausar()
        return

    filas = []
    for r in lista:
        cli = db.obtener_cliente(r.id_cliente)
        sala = db.obtener_sala(r.id_sala)
        filas.append([
            str(r.folio),
            r.evento,
            f"{cli.apellidos}, {cli.nombres}" if cli else r.id_cliente,
            sala.nombre if sala else r.id_sala,
            TURNOS.get(r.turno, r.turno),
            str(sala.cupo if sala else ""),
        ])

    print(f"\n╔{'═' * 78}╗")
    print(f"║  RESERVACIONES DEL {fecha_a_str(fecha_dt)}".ljust(79) + "║")
    print(f"╚{'═' * 78}╝")
    print(tabla(["Folio", "Evento", "Cliente", "Sala", "Turno", "Cupo"], filas))

    print("\n" + linea())
    print("¿Desea exportar el reporte?")
    print("  1) CSV")
    print("  2) JSON")
    print("  3) Excel (XLSX)")
    print("  0) No exportar")
    print(linea())
    export_op = input("Seleccione una opción: ").strip()

    if export_op in {"1", "2", "3"}:
        export_dir = "exportaciones"
        os.makedirs(export_dir, exist_ok=True)
        nombre_base = os.path.join(export_dir, f"reporte_{fecha_a_str(fecha_dt).replace('-', '')}")
        try:
            if export_op == "1":
                with open(nombre_base + ".csv", "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["Folio", "Evento", "Cliente", "Sala", "Turno", "Cupo"])
                    writer.writerows(filas)
                print(f"\n✓ Reporte exportado como {nombre_base}.csv")

            elif export_op == "2":
                data = [
                    {"Folio": f[0], "Evento": f[1], "Cliente": f[2], "Sala": f[3], "Turno": f[4], "Cupo": f[5]}
                    for f in filas
                ]
                with open(nombre_base + ".json", "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"\n✓ Reporte exportado como {nombre_base}.json")

            elif export_op == "3":
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, Border, Side

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Reservaciones"

                    ws.merge_cells('A1:F1')
                    titulo_cell = ws['A1']
                    titulo_cell.value = f"RESERVACIONES DEL {fecha_a_str(fecha_dt)}"
                    titulo_cell.font = Font(bold=True, size=14)
                    titulo_cell.alignment = Alignment(horizontal="center")

                    headers = ["Folio", "Evento", "Cliente", "Sala", "Turno", "Cupo"]
                    ws.append(headers)

                    bold = Font(bold=True)
                    border_grueso = Border(bottom=Side(border_style="thick"))
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=2, column=col)
                        cell.font = bold
                        cell.border = border_grueso
                        cell.alignment = Alignment(horizontal="center")

                    for fila in filas:
                        ws.append(fila)

                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                        for cell in row:
                            cell.alignment = Alignment(horizontal="center")

                    for col in ws.columns:
                        max_len = max(len(str(c.value)) for c in col if c.value)
                        ws.column_dimensions[col[0].column_letter].width = max_len + 2

                    wb.save(nombre_base + ".xlsx")
                    print(f"\n✓ Reporte exportado como {nombre_base}.xlsx")
                except ImportError:
                    print("\n✗ Error: El módulo 'openpyxl' no está instalado.")
                    print("   Instálelo con: pip install openpyxl")

        except Exception as e:
            print(f"\n✗ Error al exportar: {e}")
    else:
        print("\nNo se exportó el reporte.")
    pausar()

def opcion_cancelar_reservacion(db: BaseDatos):
    print(linea())
    print("CANCELAR UNA RESERVACIÓN")
    print(linea())

    fecha_desde = input_fecha("Fecha inicial del rango (mm-dd-aaaa): ")
    fecha_hasta = input_fecha("Fecha final del rango (mm-dd-aaaa): ")

    try:
        reservas = db.reservas_en_rango(fecha_desde, fecha_hasta)

        if not reservas:
            print(f"\n⚠ No hay reservaciones activas en el rango {fecha_a_str(fecha_desde)} a {fecha_a_str(fecha_hasta)}")
            pausar()
            return

        print(f"\nReservaciones del {fecha_a_str(fecha_desde)} al {fecha_a_str(fecha_hasta)}:")
        print(linea())
        filas = [[str(r.folio), r.evento, fecha_a_str(r.fecha)] for r in reservas]
        print(tabla(["Folio", "Nombre del Evento", "Fecha"], filas))
        print(linea())

        while True:
            folio_str = input("\nFolio de la reservación a cancelar (o 'CANCELAR' para salir): ").strip()
            if folio_str.upper() == 'CANCELAR':
                print("Operación cancelada.")
                pausar()
                return
            try:
                folio = int(folio_str)
                if folio in [r.folio for r in reservas]:
                    break
            except ValueError:
                pass
            print("\n⚠ El folio indicado no pertenece a este rango.")

        reserva = next(r for r in reservas if r.folio == folio)
        cliente = db.obtener_cliente(reserva.id_cliente)
        sala = db.obtener_sala(reserva.id_sala)

        print("\n" + linea())
        print("DETALLES DE LA RESERVACIÓN A CANCELAR:")
        print(linea())
        print(f"  Folio:   {reserva.folio}")
        print(f"  Evento:  {reserva.evento}")
        print(f"  Cliente: {cliente.apellidos}, {cliente.nombres}" if cliente else reserva.id_cliente)
        print(f"  Sala:    {sala.nombre}" if sala else reserva.id_sala)
        print(f"  Fecha:   {fecha_a_str(reserva.fecha)}")
        print(f"  Turno:   {TURNOS.get(reserva.turno, reserva.turno)}")
        print(linea())

        confirmacion = input("\n¿Está seguro que desea cancelar esta reservación? (S/N): ").strip().upper()

        if confirmacion != 'S':
            print("Cancelación abortada.")
            pausar()
            return

        db.cancelar_reservacion(folio)

        print("\n" + linea())
        print("✓ RESERVACIÓN CANCELADA EXITOSAMENTE")
        print("  La disponibilidad de la sala ha sido recuperada.")
        print(linea())

    except ValueError as e:
        print(f"\n✗ Error: {e}")
    except sqlite3.Error as e:
        print(f"\n✗ Error de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    pausar()

def menu():
    print("\n" + "=" * 60)
    print("SISTEMA DE RESERVACIONES DE ESPACIOS DE COWORKING")
    print("=" * 60)

    existe_bd = os.path.exists(DB_FILE)
    if existe_bd:
        print("\n>>> Se encontró una versión anterior del estado.")
    else:
        print("\n>>> No se encontró una versión anterior del estado.")
        print(">>> Se inicia con un estado inicial vacío.")

    db = BaseDatos()

    pausar()

    opciones = {
        "1": ("Registrar la reservación de una sala", opcion_registrar_reserva),
        "2": ("Editar el nombre del evento de una reservación", opcion_editar_evento),
        "3": ("Consultar las reservaciones existentes para una fecha específica", opcion_consultar_por_fecha),
        "4": ("Cancelar una reservación", opcion_cancelar_reservacion),
        "5": ("Registrar a un nuevo cliente", opcion_registrar_cliente),
        "6": ("Registrar una sala", opcion_registrar_sala),
        "7": ("Salir", None),
    }

    try:
        while True:
            os.system("cls" if os.name == "nt" else "clear")
            print("=" * 60)
            print("MENÚ PRINCIPAL - SISTEMA DE RESERVACIONES COWORKING")
            print("=" * 60)
            for k in sorted(opciones.keys()):
                print(f"  {k}. {opciones[k][0]}")
            print("=" * 60)
            op = input("Seleccione una opción: ").strip()

            if op == "7":
                print("\n" + linea())
                print("¿Está seguro que desea salir del sistema?")
                confirm = input("Confirmar salida (S/N): ").strip().upper()
                if confirm == "S":
                    print("\n✓ Saliendo del sistema.")
                    print("   (El estado se mantiene en la base de datos)")
                    print("¡Hasta luego!\n")
                    break
                else:
                    print("\nOperación cancelada. Regresando al menú principal.")
                    pausar()
            elif op in opciones:
                _, fn = opciones[op]
                if fn:
                    os.system("cls" if os.name == "nt" else "clear")
                    fn(db)
            else:
                print("\n⚠ Opción inválida.")
                pausar()
    except sqlite3.Error as e:
        print(f"\n✗ Error crítico de base de datos: {e}")
    except Exception:
        print(f"\n✗ Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        db.cerrar()

if _name_ == "_main_":
    menu()